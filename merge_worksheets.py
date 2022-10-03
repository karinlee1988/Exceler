#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/10/4 0:00
# @Author : karinlee
# @FileName : merge_worksheets.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import os
import openpyxl

from openpyxl.utils import column_index_from_string  # ,get_column_letter

class MergeXlsxWorkSheets(object):
    """
    将某个文件夹下面所有的.xlsx文件中某个sheet页的内容合并,生成一个工作薄
        仅支持.xlsx格式
        合并后的工作薄第一列的内容会指明该行内容来源为哪个.xlsx文件

    20210808 test OK

    """

    def __init__(self,folder_path:str):
        """
        :param folder_path: 待合并工作薄文件夹路径
        :type folder_path: str
        """
        self.folder_path = folder_path

    def get_singlefolder_xlsx_fullfilename(self) -> list:
        """
        获取待合并工作薄文件夹路径里指定后缀的文件名（单个文件夹，不包括子文件夹的文件）


        :return: 文件夹里面所有全文件名列表（包含子文件夹里面的文件）
        :rtype:  list
        """
        filename_list = []
        files = os.listdir(self.folder_path)
        for file in files:
            # os.path.splitext():分离文件名与扩展名
            if os.path.splitext(file)[1] in ['.xlsx']:
                filename_list.append(self.folder_path + '\\' + file)
        return filename_list

    def merge_xlsx_workbooks(self,sheet_index:str or int, title_row:str or int ,solid_column:str) -> None:
        """
        合并主程序入口

        :param sheet_index: 待合并的sheet页(从1开始)
        :type sheet_index: str or int

        :param title_row: 表头行数（从1开始）
        :type title_row: str or int

        :param solid_column: 按照某列存在的数据进行合并，避免合并空行(列号)
        :type solid_column: str

        :return: None
        :rtype:  None
        """

        # 转换为int，确保后续处理没问题
        counter = 0
        sheet_index = int(sheet_index)
        title_row = int(title_row)
        # 获取path路径下所有xlsx文件的文件名（文件名包含路径）
        xlsx_filename_list = self.get_singlefolder_xlsx_fullfilename()
        # 新建合并后的工作薄
        main_workbook = openpyxl.Workbook()
        main_worksheet = main_workbook.active
        # 构建表头 表头数据随便从文件列表中第一个文件处获取
        temp_workbook = openpyxl.load_workbook(xlsx_filename_list[0])
        temp_worksheet = temp_workbook[temp_workbook.sheetnames[sheet_index - 1]]
        list_all_title = []
        list_row_title = []
        for each_title_row in range(1,title_row+1):
            for cell in temp_worksheet[each_title_row]:
                list_row_title.append(cell.value)
            list_all_title.append(list_row_title)
            list_row_title = []
        for title in list_all_title:
            # 在最左边插入空列，便于后续处理
            title.insert(0,'')
            main_worksheet.append(title)
        # 构建数据部分，需要从每个文件中拿取
        for filename in xlsx_filename_list:
            merge_workbook = openpyxl.load_workbook(filename)
            merge_worksheet = merge_workbook[merge_workbook.sheetnames[sheet_index - 1]]
            list_all = []
            list_row = []
            for row in range(title_row+1, merge_worksheet.max_row + 1):
                # 合并工作薄的表格中，某些行可能会有空数据，使用solid_column变量确保某列存在数据的才被合并进去
                #-------------------------------------------------------------------------------------
                # 这里也可以改为其他条件，确保符合条件的数据行才会被提取并写入到合并后的工作薄
                if merge_worksheet.cell(row=row,column=column_index_from_string(solid_column)).value:
                #-------------------------------------------------------------------------------------
                    for cell in merge_worksheet[row]:  # 对当前行遍历所有单元格
                        list_row.append(cell.value)  # list_row是临时的1维列表 在遍历单元格获得每个单元格的值后写入列表 从而存储当前行的数据
                    list_all.append(list_row)  # list_all是二维列表 里面的每个元素都是1个list_row
                    list_row = []  # 重新初始化list_row列表
            for row in list_all:  # 对于2维列表的每个元素（每个元素就是每个1维列表 这些1维列表就等于excel表中1行的数据
                # 在最左边插入一列，内容为合并前的文件来源
                row.insert(0,os.path.basename(filename))
                main_worksheet.append(row)  # 可以直接使用worksheet.append()方法写入工作表中
            counter += 1
            print(f">>>{counter}<<<  {os.path.basename(filename)}  已写入合并工作薄中！")

        main_workbook.save("合并完成.xlsx")

if __name__ == '__main__':
    merger = MergeXlsxWorkSheets('tests\\test_merge_worksheets')
    merger.merge_xlsx_workbooks(sheet_index=1,title_row=3,solid_column='C')