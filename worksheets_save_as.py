#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/9/28 22:19
# @Author : karinlee
# @FileName : worksheets_save_as.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import openpyxl

def worksheet_save_as(workbook_fullfilename:str) -> None:
    """
    将一个工作薄里面的多个工作表分别另存为独立的工作薄，独立的工作薄名称为原工作薄各工作表表名
    注意：拆分后表格格式可能有所变化

    20220928 test OK

    :param workbook_fullfilename:需要进行工作表另存为的.xlsx文件全文件名
    :type workbook_fullfilename: str

    :return: None

    """
    workbook = openpyxl.load_workbook(workbook_fullfilename)
    sheetname_list = workbook.sheetnames
    for name in sheetname_list:
        worksheet = workbook[name]
        # 创建新的Excel
        workbook_new = openpyxl.Workbook()
        # 获取当前sheet
        worksheet_new = workbook_new.active
        # 两个for循环遍历整个excel的单元格内容
        for i, row in enumerate(worksheet.iter_rows(),start=1): #enumerate()从1开始  # 或者for i, row in enumerate(worksheet.rows):
            for j, cell in enumerate(row,start=1):#enumerate()从1开始
                # 写入新Excel
                worksheet_new.cell(row=i, column=j, value=cell.value)
                # 设置新Sheet的名称
                worksheet_new.title = name
        workbook_new.save(name + '.xlsx')

if __name__ == '__main__':
    wb = 'tests\\test_worksheets_save_as\\拆分工作表测试.xlsx'
    worksheet_save_as(wb)