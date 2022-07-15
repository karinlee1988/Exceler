#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/7/12 20:04
# @Author : karinlee
# @FileName : easy_vlookup.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/


import os
import openpyxl
import tkinter as tk
import tkinter.filedialog
from openpyxl.utils import column_index_from_string  # ,get_column_letter

def easyvlookup(
        wb_main,
        ws_main_index:int,
        main_key:str,
        main_value:str,
        wb_source,
        ws_source_index:int,
        source_key:str,
        source_value:str,
        line:int
        ) -> None:
    """
    对2个不同的工作薄执行vlookup操作

    主工作薄：需要写入数据的工作薄
    数据工作薄：根据模板工作薄提供的条件（列）在数据工作薄中查找，提供数据来源的工作薄
    注意！函数执行完后，只对wb_main对象进行了数据写入。在函数外部还需wb_main.save("filename.xlsx"),vlookup后的数据才能保存为excel表。

    20220714 test OK

    :param:
        'wb_main': 主工作薄对象
        'ws_main_index': 需要处理的主工作表索引号
        'main_key': 主工作表key所在列号
        'main_value': 主工作表value需要填写的列号
        'wb_source': 数据工作薄对象
        'ws_source_index': 需要处理的数据工作表索引号
        'source_key':  数据工作表key所在列号
        'source_value': 数据工作表value所在列号
        'line' :从第几行开始vlookup

    :type:
        'wb_main': class Workbook
        'ws_main_index': int
        'main_key': str
        'main_value': str
        'wb_source': class Workbook
        'ws_source_index': int
        'source_key':  str
        'source_value': str
        'line' :int

    :return: None

    """
    # 获取数据工作表对象
    ws_source = wb_source[wb_source.sheetnames[ws_source_index-1]]
    # 获取主工作表对象
    ws_main = wb_main[wb_main.sheetnames[ws_main_index-1]]
    # 获取数据工作表的查找列和数据列，分别生成2个元组
    source_key_tuple = ws_source[source_key]
    source_value_tuple = ws_source[source_value]
    # 创建2个列表，遍历元组，将元组中每个单元格的值添加到列表中
    list_key =[]
    list_value = []
    for cell in source_key_tuple:
        list_key.append(cell.value)
    for cell in source_value_tuple:
        list_value.append(cell.value)
    # 通过数据工作表的key列和value列，创建好需要进行vlookup的字典
    dic = dict(zip(list_key,list_value))
    # 将列号 （str）转为列索引值 （int）
    main_key_index = column_index_from_string(main_key)
    main_value_index = column_index_from_string(main_value)
    # 从第line行开始进行vlookup  可根据表头行数进行修改
    for row in range(int(line),ws_main.max_row+1):
        # ------------------------------------------------------
        # 采用dict[key]方式查字典，如果没有key的话会raise KeyError
        # try:
        #     ws_main.cell(row=row,column=template_value_index).value = dic[
        #     ws_main.cell(row=row,column=template_key_index).value]
        # except KeyError:
        #     #找不到数据 相应的单元格填上#N/A
        #     ws_main.cell(row=row, column=template_value_index).value = "#N/A"
        # -------------------------------------------------------
        #采用dict.get()避免出现keyerror
        ws_main.cell(row=row, column=main_value_index).value = dic.get(
            ws_main.cell(row=row, column=main_key_index).value)

class EasyVlookupGui(object):
    """
    python实现vlookup图形工具

    20220714 test OK
    """
    def __init__(self):
        """
        创建界面
        """
        # 新建窗口
        self.master = tk.Tk()
        # 在界面顶部添加横幅图片
        self.photo = tk.PhotoImage(file="images\\doge2.gif")
        # self.path 用于存放选择的文件路径
        # self.flag 当程序运行完成后给用户提示信息
        # 注意！这些都是tk.StringVar()对象，不是str，其他地方要用的话要用get()方法获取str
        self.mainbook_path = tk.StringVar()
        self.sourcebook_path = tk.StringVar()
        self.flag = tk.StringVar()
        self.v1 = tk.StringVar()  # 主工作薄工作表序号
        self.v2 = tk.StringVar()   # 主工作薄用于查找的列
        self.v3 = tk.StringVar()   # 主工作薄要填写的列
        self.v4 = tk.StringVar()   # 数据工作薄工作表序号
        self.v5 = tk.StringVar()  # 数据工作薄对应主工作薄的列
        self.v6 = tk.StringVar()  # 数据工作薄提供数据列
        self.v7 = tk.StringVar()  # 从第几行开始vlookup
        # 窗口图片横幅
        self.img_pack()
        # 主界面布局
        self.window()
        # 框架布局
        self.frame()
        # 保持运行
        self.master.mainloop()

    def img_pack(self):
        """
        窗口图片横幅
        """
        # 图片贴上去
        img_lable = tk.Label(self.master, image=self.photo)
        img_lable.pack()

    def window(self):
        """
        主界面布局设置
        """
        # 调整窗口默认大小及在屏幕上的位置
        self.master.geometry("800x820+550+20")
        # 窗口的标题栏，自己修改
        self.master.title("  EasyVlookup  by李加林v2.0")
        # 把标题贴上去，自己修改
        tk.Label(self.master,text="EasyVlookup",font=("黑体",18)).pack()
        tk.Label(self.master, text="-----------------------", font=("黑体", 16)).pack()

    def frame(self):
        """
        框架布局设置

        """
        # 框架贴上去，再在框架里添加Lable，Entry，Button等控件
        frame1 = tk.Frame(self.master)
        frame1.pack()
        frame2 = tk.Frame(self.master)
        frame2.pack()
        # 输入框，标记，按键
        tk.Label(frame1, text="目标路径:", font=("黑体", 14)).grid(row=1, column=0)
        tk.Label(frame1, text="主工作薄->", font=("黑体", 14)).grid(row=2, column=0)
        tk.Entry(frame1, textvariable=self.mainbook_path, width=50).grid(row=2, column=1)
        tk.Label(frame1, text="数据工作薄->", font=("黑体", 14)).grid(row=3, column=0)
        tk.Entry(frame1, textvariable=self.sourcebook_path, width=50).grid(row=3, column=1)
        tk.Button(frame1, text="选择主工作薄", command=self.select_mainbook_path, font=("黑体", 14)).grid(row=2, column=2)
        tk.Button(frame1, text="选择数据工作薄", command=self.select_source_path, font=("黑体", 14)).grid(row=3, column=2)

        tk.Label(frame2, text="主工作薄工作表序号（从1开始）:", font=("黑体", 14)).grid(row=1, column=0)
        tk.Entry(frame2, textvariable=self.v1).grid(row=1, column=1)
        tk.Label(frame2, text="主工作薄用于查找的列名(大写字母):", font=("黑体", 14)).grid(row=2, column=0)
        tk.Entry(frame2, textvariable=self.v2).grid(row=2, column=1)
        tk.Label(frame2, text="主工作薄要填写的列名(大写字母):", font=("黑体", 14)).grid(row=3, column=0)
        tk.Entry(frame2, textvariable=self.v3).grid(row=3 ,column=1)
        tk.Label(frame2, text="数据工作薄工作表序号（从1开始）:", font=("黑体", 14)).grid(row=4, column=0)
        tk.Entry(frame2, textvariable=self.v4).grid(row=4, column=1)
        tk.Label(frame2, text="数据工作薄对应主工作薄的列名(大写字母):", font=("黑体", 14)).grid(row=5, column=0)
        tk.Entry(frame2, textvariable=self.v5).grid(row=5, column=1)
        tk.Label(frame2, text="数据工作薄提供数据列名(大写字母):", font=("黑体", 14)).grid(row=6, column=0)
        tk.Entry(frame2, textvariable=self.v6).grid(row=6 ,column=1)
        tk.Label(frame2, text="从第几行开始需要vlookup:", font=("黑体", 14)).grid(row=7, column=0)
        tk.Entry(frame2, textvariable=self.v7).grid(row=7, column=1)

        # 按这个按钮执行主程序
        tk.Button(frame2, text="开始VLOOKUP", command=self.main, font=("黑体", 14)).grid(row=8, column=1,pady=33)
        tk.Entry(frame2, textvariable=self.flag,state="readonly").grid(row=9, column=1)

    def select_mainbook_path(self):
        """
        选择文件，并获取文件的绝对路径

        """
        # 选择文件，path_select变量接收文件地址
        # 注意：self.path 是tk.StringVar()对象，而path_select是str变量
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.mainbook_path.set(path_select)
        self.flag.set("准备进行vlookup...")

    def select_source_path(self):
        """
        选择文件，并获取文件的绝对路径

        """
        # 选择文件，path_select变量接收文件地址
        # 注意：self.path 是tk.StringVar()对象，而path_select是str变量
        path_select = tkinter.filedialog.askopenfilename()
        # 通过replace函数替换绝对文件地址中的/来使文件可被程序读取
        # 注意：\\转义后为\，所以\\\\转义后为\\
        path_select = path_select.replace("/", "\\\\")
        # self.path设置path_select的值
        self.sourcebook_path.set(path_select)

    def main(self):
        """
        程序入口
        """
        wb_main = openpyxl.load_workbook(self.mainbook_path.get())
        wb_source = openpyxl.load_workbook(self.sourcebook_path.get())

        ws_main_index = int(self.v1.get())
        main_key = self.v2.get()
        main_value = self.v3.get()

        ws_source_index = int(self.v4.get())
        source_key = self.v5.get()
        source_value = self.v6.get()
        line = int(self.v7.get())
        # 进行vlookup处理
        easyvlookup(wb_main=wb_main,
                    ws_main_index=ws_main_index,
                    main_key=main_key,
                    main_value=main_value,
                    wb_source=wb_source,
                    ws_source_index=ws_source_index,
                    source_key=source_key,
                    source_value=source_value,
                    line=line)
        wb_main.save(os.path.basename(self.mainbook_path.get()).replace('.xlsx', '') + '_已进行vlookup.xlsx')
        # 标志设置为处理完成
        self.flag.set("处理完成！")

if __name__ == '__main__':
    # 以下调用easyvlookup()函数 例1
    # wbmain = openpyxl.load_workbook('tests\\tests_easy_vlookup\\主表.xlsx')
    # wbsource = openpyxl.load_workbook('tests\\tests_easy_vlookup\\数据表.xlsx')
    # easyvlookup(wb_main=wbmain,
    #             ws_main_index=1,
    #             main_key='B',
    #             main_value='D',
    #             wb_source=wbsource,
    #             ws_source_index=2,
    #             source_key='K',
    #             source_value='J',
    #             line=2)
    # wbmain.save('已进行vlookup.xlsx')
    #--------------------------------------------
    ## 以下调用easyvlookup()函数 例2
    # wbmain = openpyxl.load_workbook('tests\\tests_easy_vlookup\\20220615_2021年度失业保险支持企业稳定岗位返名单（大型企业返还比例从30%转为50%）(添加银行信息).xlsx')
    # wbsource = openpyxl.load_workbook('tests\\tests_easy_vlookup\\全市2021年度失业保险稳岗返还企业数据 20220418整理(市局王岳岳20220419提供).xlsx')
    # easyvlookup(wb_main=wbmain,
    #             ws_main_index=1,
    #             main_key='D',
    #             main_value='R',
    #             wb_source=wbsource,
    #             ws_source_index=3,
    #             source_key='C',
    #             source_value='L',
    #             line=2)
    # wbmain.save('已进行vlookup2.xlsx')
    # -------------------------------------------
    # 以下运行图形界面
    app = EasyVlookupGui()

