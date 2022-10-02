#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/10/2 23:09
# @Author : karinlee
# @FileName : worksheet_split_regular_line.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import os
import openpyxl

class WorksheetSplitRegularLine(object):
    """
    按照固定行数将单个工作表拆分为多个工作表(如将1个3000行的表格拆分为3个1000行的表格)
    表头默认为1行
    对于工作薄，操作对象默认为工作薄的第一个工作表

    20221002 test OK
    """

    def __init__(self,full_filename:str,batche_num:int or str):
        """

        :param full_filename: 待拆分xlsx文件的全文件名（相对路径或绝对路径）
        :param batche_num: 需要拆分的固定行数

        :type full_filename: str
        :type batche_num: int or str
        """
        self.full_filename = full_filename
        self.batche_num = batche_num
        # 获取self.workbook 要拆分的工作薄对象
        self.workbook = openpyxl.load_workbook(self.full_filename)
        # 获取worksheet，默认为第一个sheet用于拆分
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 把生成器转换为列表
        self.lines = list(self.sheet.rows)
        # 获取第一行合并行
        # first_line = lines[0]
        # 获取表头字段行
        self.header = self.lines[0:1]
        # 获取数据行
        self.dataline = self.lines[1:]

    def one_sheet(self,data, sheet_no):
        """
        把一个数据集保存成一个xlsx文件，。
        :param data: 数据集
        :param sheet_no: 拆分文件的流水号
        :type data: list
        :type sheet_no: int
        :return: None
        """
        # 创建一个xlsx文件对象
        wb = openpyxl.Workbook()
        # 取得默认的worksheet
        ws = wb.active
        # ws.title = '新测试表%02d' % (sheet_no+1)   # 设置一个标题
        # # 第一行写入合并行
        # ws.cell(row=1, column=1).value = first_line[0].value
        # # 该行所有列合并
        # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(first_line))
        # 写头：循环写每个字段的值：行从1开始，所以表头行索引是2
        header_idx = 0
        for col in self.header[0]:
            ws.cell(row=1, column=(header_idx + 1)).value = col.value
            header_idx += 1
        # 纪录行索引
        row_idx = 0
        for row_ in data:
            col_idx = 0
            for col in row_:
                # 数据行行从2开始
                ws.cell(row=(row_idx+2) ,column=(col_idx + 1)).value = col.value
                # 纪录列索引
                col_idx += 1
            row_idx += 1
        # 保存文件
        wb.save(os.path.basename(self.full_filename)+'-%02d.xlsx' % (sheet_no + 1))

    def main(self):
        """
        程序入口，运行后开始拆分
        """
        # 计算拆分文件个数
        batches = (len(self.dataline) // self.batche_num) + 1
        # 循环写数据集到每一个文件
        for pt in range(batches):
            # 取数据集，每个数据集最多batche_num行，最后一个不足batche_num行，直接处理。
            lines_ = self.dataline[pt * self.batche_num : (pt + 1) * self.batche_num]
            self.one_sheet(lines_,pt)

if __name__ == '__main__':
    cut = WorksheetSplitRegularLine("tests\\test_worksheet_split_regular_line\\固定行数拆分测试.xlsx",100)
    cut.main()