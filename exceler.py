#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2021/8/1 15:12
# @Author : karinlee
# @FileName : exceler.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

import os
import openpyxl

from openpyxl.utils import column_index_from_string  # ,get_column_letter



def worksheet_save_as(workbook) -> None:
    """
    将一个工作薄里面的多个工作表分别另存为独立的工作薄，独立的工作薄名称为原工作薄各工作表表名

    20210808 test OK

    :param workbook:需要进行工作表另存为的workbook对象
    :type workbook: class Workbook

    :return: None

    """
    sheetname_list = workbook.sheetnames
    for name in sheetname_list:
        worksheet = workbook[name]
        # 创建新的Excel
        workbook_new = openpyxl.Workbook()
        # 获取当前sheet
        worksheet_new = workbook_new.active
        # 两个for循环遍历整个excel的单元格内容
        for i, row in enumerate(worksheet.iter_rows(),start=1): #enumerate()从1开始  # 或者for i, row in enumerate(worksheet.rows):
            for j, cell in enumerate(row,start=1):#enumerate()从1开始
                # 写入新Excel
                worksheet_new.cell(row=i, column=j, value=cell.value)
                # 设置新Sheet的名称
                worksheet_new.title = name
        workbook_new.save(name + '.xlsx')

class MergeXlsxWorkbooks(object):
    """
    将某个文件夹下面所有的.xlsx文件中某个sheet页的内容合并为一个工作薄
        仅支持.xlsx格式
        合并后的工作薄第一列的内容会指明该行内容来源为哪个.xlsx文件

    20210808 test OK

    """

    def __init__(self,folder_path:str):
        """
        :param folder_path: 待合并工作薄文件夹路径
        :type folder_path: str
        """

        self.folder_path = folder_path

    def get_singlefolder_xlsx_fullfilename(self) -> list:
        """
        获取待合并工作薄文件夹路径里指定后缀的文件名（单个文件夹，不包括子文件夹的文件）


        :return: 文件夹里面所有全文件名列表（包含子文件夹里面的文件）
        :rtype:  list
        """
        filename_list = []
        files = os.listdir(self.folder_path)
        for file in files:
            # os.path.splitext():分离文件名与扩展名
            if os.path.splitext(file)[1] in ['.xlsx']:
                filename_list.append(self.folder_path + '\\' + file)
        return filename_list


    def merge_xlsx_workbooks(self,sheet_index:str or int, title_row:str or int ,solid_column:str) -> None:
        """
        合并主程序



        :param sheet_index: 待合并的sheet页(从1开始)
        :type sheet_index: str or int

        :param title_row: 表头行数（从1开始）
        :type title_row: str or int

        :param solid_column: 按照某列存在的数据进行合并，避免合并空行(列号)
        :type solid_column: str

        :return: None
        :rtype:  None
        """

        # 转换为int，确保后续处理没问题
        counter = 0
        sheet_index = int(sheet_index)
        title_row = int(title_row)
        # 获取path路径下所有xlsx文件的文件名（文件名包含路径）
        xlsx_filename_list = self.get_singlefolder_xlsx_fullfilename()
        # 新建合并后的工作薄
        main_workbook = openpyxl.Workbook()
        main_worksheet = main_workbook.active
        # 构建表头 表头数据随便从文件列表中第一个文件处获取
        temp_workbook = openpyxl.load_workbook(xlsx_filename_list[0])
        temp_worksheet = temp_workbook[temp_workbook.sheetnames[sheet_index - 1]]
        list_all_title = []
        list_row_title = []
        for each_title_row in range(1,title_row+1):
            for cell in temp_worksheet[each_title_row]:
                list_row_title.append(cell.value)
            list_all_title.append(list_row_title)
            list_row_title = []
        for title in list_all_title:
            # 在最左边插入空列，便于后续处理
            title.insert(0,'')
            main_worksheet.append(title)
        # 构建数据部分，需要从每个文件中拿取
        for filename in xlsx_filename_list:

            merge_workbook = openpyxl.load_workbook(filename)
            merge_worksheet = merge_workbook[merge_workbook.sheetnames[sheet_index - 1]]
            list_all = []
            list_row = []
            for row in range(title_row+1, merge_worksheet.max_row + 1):
                # 合并工作薄的表格中，某些行可能会有空数据，使用solid_column变量确保某列存在数据的才被合并进去
                #-------------------------------------------------------------------------------------
                # 这里也可以改为其他条件，确保符合条件的数据行才会被提取并写入到合并后的工作薄
                if merge_worksheet.cell(row=row,column=column_index_from_string(solid_column)).value:
                #-------------------------------------------------------------------------------------
                    for cell in merge_worksheet[row]:  # 对当前行遍历所有单元格
                        list_row.append(cell.value)  # list_row是临时的1维列表 在遍历单元格获得每个单元格的值后写入列表 从而存储当前行的数据
                    list_all.append(list_row)  # list_all是二维列表 里面的每个元素都是1个list_row
                    list_row = []  # 重新初始化list_row列表
            for row in list_all:  # 对于2维列表的每个元素（每个元素就是每个1维列表 这些1维列表就等于excel表中1行的数据
                # 在最左边插入一列，内容为合并前的文件来源
                row.insert(0,os.path.basename(filename))

                main_worksheet.append(row)  # 可以直接使用worksheet.append()方法写入工作表中
            counter += 1
            print(f">>>{counter}<<<  {os.path.basename(filename)}  已写入合并工作薄中！")

        main_workbook.save("合并完成.xlsx")

class WorkbookSplitRegularLine(object):
    """
    按照固定行数将单个工作薄拆分为多个工作薄(如将1个3000行的表格拆分为3个1000行的表格)
    表头默认为1行
    """

    def __init__(self,full_filename:str,batche_num:int or str):
        """

        :param full_filename: 待拆分xlsx文件的全文件名（相对路径或绝对路径）
        :param batche_num: 需要拆分的行数

        :type full_filename: str
        :type batche_num: int or str
        """
        self.full_filename = full_filename
        self.batche_num = batche_num
        # 获取self.workbook 要拆分的工作薄对象
        self.workbook = openpyxl.load_workbook(self.full_filename)
        # 获取worksheet，默认为第一个sheet用于拆分
        self.sheet = self.workbook[self.workbook.sheetnames[0]]
        # 把生成器转换为列表
        self.lines = list(self.sheet.rows)
        # 获取第一行合并行
        # first_line = lines[0]
        # 获取表头字段行
        self.header = self.lines[0:1]
        # 获取数据行
        self.dataline = self.lines[1:]

    def one_sheet(self,data, sheet_no):
        """
        把一个数据集保存成一个xlsx文件，。
        :param data: 数据集
        :param sheet_no: 拆分文件的流水号
        :type data: list
        :type sheet_no: int
        :return: None
        """
        # 创建一个xlsx文件对象
        wb = openpyxl.Workbook()
        # 取得默认的worksheet
        ws = wb.active
        # ws.title = '新测试表%02d' % (sheet_no+1)   # 设置一个标题
        # # 第一行写入合并行
        # ws.cell(row=1, column=1).value = first_line[0].value
        # # 该行所有列合并
        # ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(first_line))
        # 写头：循环写每个字段的值：行从1开始，所以表头行索引是2
        header_idx = 0
        for col in self.header[0]:
            ws.cell(row=1, column=(header_idx + 1)).value = col.value
            header_idx += 1
        # 纪录行索引
        row_idx = 0
        for row_ in data:
            col_idx = 0
            for col in row_:
                # 数据行行从2开始
                ws.cell(row=(row_idx+2) ,column=(col_idx + 1)).value = col.value
                # 纪录列索引
                col_idx += 1
            row_idx += 1
        # 保存文件
        wb.save(os.path.basename(self.full_filename)+'-%02d.xlsx' % (sheet_no + 1))

    def dealing(self):
        # 计算拆分文件个数
        batches = (len(self.dataline) // self.batche_num) + 1
        # 循环写数据集到每一个文件
        for pt in range(batches):
            # 取数据集，每个数据集最多3000行，最后一个不足3000行，直接处理。
            lines_ = self.dataline[pt * self.batche_num : (pt + 1) * self.batche_num]
            self.one_sheet(lines_,pt)

if __name__ == '__main__':
    # mer = MergeXlsxWorkbooks(r'C:\LIJIALIN\BaiduNetdiskWorkspace\HOME\学习\PYTHON\GitHub和Gitee仓库\Exceler\tests\各单位汇总_xlsx')
    # mer.merge_xlsx_workbooks(1,3,'C')
    cut = WorkbookSplitRegularLine("tests\\固定行数拆分测试.xlsx",30)
    cut.dealing()
