#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/7/10 11:38
# @Author : karinlee
# @FileName : worksheet_split_by_column.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/

"""
用于按列的内容拆分excel .XLSX工作薄中的一个工作表，分别另存为多个独立的工作薄/打印
20220712 test OK
"""

import os
import openpyxl

class WorksheetSplitByColumn(object):
    """
    按列拆分工作表为多个独立表格/拆分时打印可选
    20220712 test OK
    """


    def __init__(self,filepath,column_index:int,title_index:int,isprint:bool=False):

        """

        :param filepath: excel xlsx文件路径
        :param column_index: 按哪列的内容进行拆分（A列为1，B列为2,C列为3...）
        :param title_index: 表头的行数（从1开始，即表头有几行就填写几行）

        """
        # 获取文件路径
        self.filepath = filepath
        # 获取作为条件拆分的列号
        self.column_index = column_index
        # 获取表头行数
        self.title_index =title_index
        # 是否选择拆分后自动打印，默认为false(否)
        self.isprint = isprint

    def split_by_column(self, column_key: str):
        """
        拆分表格并分别另存为

        在待拆分表格中，采取按列关键字删除的方式来进行拆分，避免出现格式错乱
        这样，待拆分表格先行用excel打开设置好格式，删除就不会错乱
        但删除操作耗时较大。

        :param column_key:  列关键字,用于按列拆分
        :return:
        """
        # 读取待拆分的源数据表（为了不改变原有格式，每次都读取一次比较稳妥。拆分等于就是在原表上将不要的行删除掉，格式还是原表的格式）
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        # 根据传入的列关键字，若列中与传入的关键字不同，则删除。循环完毕后，剩下就等效于拆分后的数据了
        # 删除操作会改变行对应的行号，所以从最大行开始遍历，倒过来循环。
        for row in range(ws.max_row + 1, self.title_index, -1):
            if ws.cell(row=row, column=self.column_index).value != column_key:
                ws.delete_rows(row)
        # 拆分后的表格另存为独立的xlsx文件，设置文件名为列关键字
        wb.save(f'{column_key}.xlsx')
        if self.isprint is True:
            os.startfile(f'{column_key}.xlsx','print')

    def get_column_key_list(self):
        """
        获取列关键字列表

        :return: duplicate_removal_list: 去重后的列关键字列表
        :rtype: duplicate_removal_list: str

        """
        column_key_list = []
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        for row in range (self.title_index+1,ws.max_row+1):
            column_key_list.append(ws.cell(row=row,column=self.column_index).value)
        #获取到的column_key_list有重复，需要去重
        duplicate_removal_list = []
        for column_key in column_key_list:
            if column_key not in duplicate_removal_list:
                duplicate_removal_list.append(column_key)

        # 返回去重后的列关键字列表
        return duplicate_removal_list

    def main(self):
        """
        主程序入口
        :return:
        """
        # 先拿到去重的列关键字
        duplicate_removal_list = self.get_column_key_list()
        # 然后不断循环拆分
        # count变量用于计数
        count = 1
        for column_key in duplicate_removal_list:
            self.split_by_column(column_key=column_key)
            print(f'{count}  ----  {column_key}  已成功拆分！')
            #每拆分1个，计数+1
            count += 1

if __name__ == '__main__':
    workbook_path = 'tests\\test_worksheet_split_by_column\\按列拆分测试表.xlsx'
    app = WorksheetSplitByColumn(filepath=workbook_path,column_index=4,title_index=4,isprint=False)
    app.main()






